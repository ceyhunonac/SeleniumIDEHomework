from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
from constants import globalConstants
from pathlib import Path
from datetime import date
import pytest
import openpyxl


class TestShoppingtest():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.get(globalConstants.URL)
    self.driver.maximize_window()

    self.folderPath =("SeleniumIDEHomework/TestScreenshots")
    
  def teardown_method(self):
    self.driver.quit()

  def getData():
      excelFile = openpyxl.load_workbook("SeleniumIDEHomework/data.xlsx")
      selectedSheet = excelFile["Sheet1"]

      totalRows = selectedSheet.max_row
      data=[]
      for i in range(2,totalRows+1): 
          username = selectedSheet.cell(i,1).value
          password = selectedSheet.cell(i,2).value
          firstName = selectedSheet.cell(i,3).value
          lastName = selectedSheet.cell(i,4).value
          postalCode = selectedSheet.cell(i,5).value
          tupleData = (username,password,firstName,lastName,postalCode)
          data.append(tupleData)

      return data

  @pytest.mark.parametrize("username,password,firstName,lastName,postalCode",getData())
  def test_shoppingtest(self,username,password,firstName,lastName,postalCode):
    
        
    self.waitForElementVisible((By.ID,"user-name"))
    usernameInput = self.driver.find_element(By.ID,"user-name")
    usernameInput.send_keys(username)

    self.waitForElementVisible((By.ID,"password"),2)
    usernameInput = self.driver.find_element(By.ID,"password")
    usernameInput.send_keys(password)

    self.waitForElementVisible((By.ID,"login-button"),2)
    loginButton = self.driver.find_element(By.ID,"login-button")
    loginButton.click()

    self.driver.save_screenshot(f"{self.folderPath}/test-shopping-login1.0-{username}-{firstName}.png")


    self.waitForElementVisible((By.ID,"add-to-cart-sauce-labs-backpack"))
    addToCartButton = self.driver.find_element(By.ID,"add-to-cart-sauce-labs-backpack")    
    addToCartButton.click()

    self.driver.save_screenshot(f"{self.folderPath}/test-shopping-login1.1-{username}-{firstName}.png")

    self.driver.find_element(By.ID,"shopping_cart_container").click()

    self.waitForElementVisible((By.ID,"checkout"))
    checkout = self.driver.find_element(By.ID,"checkout")
    checkout.click()

    self.waitForElementVisible((By.ID,"first-name"))
    firstNameInput = self.driver.find_element(By.ID,"first-name")
    firstNameInput.send_keys(firstName)

    self.waitForElementVisible((By.ID,"last-name"),2)
    lastNameInput = self.driver.find_element(By.ID,"last-name")
    lastNameInput.send_keys(lastName)

    self.waitForElementVisible((By.ID,"postal-code"),2)
    postalCodeInput = self.driver.find_element(By.ID,"postal-code")
    postalCodeInput.send_keys(postalCode)
    
    self.driver.save_screenshot(f"{self.folderPath}/test-shopping-login1.2-{username}-{firstName}.png")
    
    self.waitForElementVisible((By.ID,"continue"),2)
    continueButton = self.driver.find_element(By.ID,"continue")
    continueButton.click()

    self.driver.save_screenshot(f"{self.folderPath}/test-shopping-login1.3-{username}-{firstName}.png")
   

    self.waitForElementVisible((By.ID,"finish"))
    finishButton = self.driver.find_element(By.ID,"finish")
    finishButton.click()

    self.waitForElementVisible((By.ID,"back-to-products"))
    backToProducts = self.driver.find_element(By.ID,"back-to-products")
    backToProducts.click()

    self.driver.save_screenshot(f"{self.folderPath}/test-shopping-login1.4-{username}-{firstName}.png")
  
  def waitForElementVisible(self,locator,timeout=3):
    WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
